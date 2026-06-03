import NXOpen
import NXOpen.UF
import NXOpen.CAM
import os
import threading
import time
import ctypes
import datetime
import traceback
import json
import re
# import openpyxl  # 用于Excel生成（NX 内置 Python 没装这个，注释掉跳过 Excel）

# ==================================================================================
# 1. 工具函数：部件检查、保存
# ==================================================================================
def check_part_load_status(load_status):


    """检查部件加载状态"""
    try:
        if load_status == 0:
            return True
    except:
        pass
    try:
        session = NXOpen.Session.GetSession()
        if session.Parts.Work is not None:
            return True
    except:
        pass
    return False

def save_part(new_path: str) -> bool:
    """保存部件：指定路径则另存"""
    session = NXOpen.Session.GetSession()
    work_part = session.Parts.Work
    if work_part is None:
        return False
    try:
        # 确保保存路径的目录存在（兜底处理）
        new_path_dir = os.path.dirname(new_path)
        if not os.path.exists(new_path_dir):
            os.makedirs(new_path_dir)
        work_part.SaveAs(new_path)
        return True
    except Exception as e:
        print(f"保存部件失败: {str(e)}")
        return False

# ==================================================================================
# 2. 刀轨生成器类
# ==================================================================================
class ToolpathGeneratorMacro:
    """刀轨生成器 - 为CAM操作生成刀轨"""
    def __init__(self, session, work_part):
        self.session = session
        self.work_part = work_part
        self.success_count = 0
        self.failed_count = 0
        self.test_results = []

    def print_log(self, message, level="INFO"):
        """打印带时间戳的日志（控制台输出，替代ListingWindow）"""
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        level_symbols = {
            "INFO": "[i]", "SUCCESS": "[OK]", "ERROR": "[ERR]",
            "WARN": "[!]", "DEBUG": "[?]", "START": "[>]", "END": "[X]"
        }
        symbol = level_symbols.get(level, "*")
        log_msg = f"[{timestamp}] {symbol} {message}"
        print(log_msg, flush=True)

    def print_separator(self, char="=", length=60):
        sep = char * length
        print(sep, flush=True)

    def print_header(self, title):
        self.print_separator()
        self.print_log(f"  {title}", "START")
        self.print_separator()

    def switch_to_manufacturing(self):
        try:
            self.session.ApplicationSwitchImmediate("UG_APP_MANUFACTURING")
            self.print_log("已切换到加工环境", "SUCCESS")
        except Exception as e:
            self.print_log(f"切换加工环境警告: {e}", "WARN")

    def get_all_operations(self):
        operations = []
        try:
            if self.work_part.CAMSetup is None:
                self.print_log("当前部件无CAM加工环境", "WARN")
                return operations
            for operation in self.work_part.CAMSetup.CAMOperationCollection:
                operations.append(operation)
            self.print_log(f"找到 {len(operations)} 个CAM操作", "INFO")
        except Exception as e:
            self.print_log(f"获取操作列表失败: {e}", "ERROR")
        return operations

    def generate_toolpath(self, operation):
        op_name = operation.Name
        try:
            # 设置Undo标记，虽然后面不一定会撤销，但由于API要求
            mark_id = self.session.SetUndoMark(
                NXOpen.Session.MarkVisibility.Visible,
                f"Generate Tool Path - {op_name}"
            )
            objects = [NXOpen.CAM.CAMObject.Null] * 1
            objects[0] = operation
            self.work_part.CAMSetup.GenerateToolPath(objects)
            self.print_log(f"刀轨生成完成: {op_name}", "SUCCESS")
            self.success_count += 1
            self.test_results.append({
                "name": op_name, "status": "Success", "message": "刀轨生成成功"
            })
            return True
        except Exception as e:
            err_msg = f"刀轨生成失败 - {op_name}: {e}"
            self.print_log(err_msg, "ERROR")
            self.failed_count += 1
            self.test_results.append({
                "name": op_name, "status": "Failed", "error": str(e)
            })
            return False

    def generate_all_toolpaths(self):
        self.print_header("NX CAM 刀轨生成流程")
        self.switch_to_manufacturing()
        operations = self.get_all_operations()

        if not operations:
            self.print_log("没有找到任何CAM操作", "WARN")
            return

        # 列出所有操作
        for i, op in enumerate(operations, 1):
            self.print_log(f"  {i}. {op.Name}")

        self.print_separator("-")
        self.print_log("开始生成刀轨...", "START")
        for operation in operations:
            self.generate_toolpath(operation)

        self.print_summary()

    def print_summary(self):
        total = self.success_count + self.failed_count
        success_rate = (self.success_count / total * 100) if total > 0 else 0
        self.print_separator("=")
        summary = f"""
  刀轨生成汇总
  ----------------------------------------
  总操作数:   {total}
  成功:       {self.success_count} [OK]
  失败:       {self.failed_count} [FAIL]
  成功率:     {success_rate:.1f}%
        """.strip()
        print(summary, flush=True)
        self.print_separator("=")

# ==================================================================================
# 3. 加特林线程类 (模拟按键处理弹窗)
# ==================================================================================
class EnterSpammer(threading.Thread):
    def __init__(self, key_interval=0.3):
        super().__init__()
        self._stop_event = threading.Event()
        self.daemon = True
        self.key_interval = key_interval

    def run(self):
        user32 = ctypes.windll.user32
        VK_RETURN = 0x0D
        counter = 0
        while not self._stop_event.is_set():
            try:
                user32.keybd_event(VK_RETURN, 0, 0, 0)
                user32.keybd_event(VK_RETURN, 0, 2, 0)
                counter += 1
                time.sleep(self.key_interval)
            except Exception as e:
                # 捕获按键异常，避免线程崩溃
                time.sleep(self.key_interval)
                continue

    def stop(self):
        self._stop_event.set()

# ==================================================================================
# 4. 过切检查核心流程（保留文件输出，移除窗口交互）
# ==================================================================================
def process_gauge_check(config):
    """执行过切检查流程"""
    session = NXOpen.Session.GetSession()
    workPart = session.Parts.Work
    if workPart is None:
        print("过切检查失败：当前无工作部件")
        return False

    # 切换加工模块
    try:
        session.ApplicationSwitchImmediate("UG_APP_MANUFACTURING")
        uf = NXOpen.UF.UFSession.GetUFSession()
        uf.Cam.InitSession()
    except Exception as e:
        print(f"切换加工模块失败，但继续执行过切检查: {e}")

    # 初始化输出窗口（仅用于写入TXT文件，无窗口交互）
    list_window = session.ListingWindow
    try:
        # 确保TXT文件目录存在（兜底处理）
        txt_dir = os.path.dirname(config['OUTPUT_PATH'])
        if not os.path.exists(txt_dir):
            os.makedirs(txt_dir)
        list_window.SelectDevice(NXOpen.ListingWindow.DeviceType.File, config['OUTPUT_PATH'])
        list_window.Open()
    except Exception as e:
        print(f"初始化输出文件失败: {e}")
        return False

    infoTool = session.Information
    camSetup = workPart.CAMSetup
    if camSetup is None:
        print("过切检查失败：当前部件无CAMSetup")
        list_window.Close()
        return False

    # 获取工序列表
    try:
        operations = [op for op in camSetup.CAMOperationCollection]
    except Exception as e:
        print(f"获取工序列表失败: {e}")
        list_window.Close()
        return False

    if not operations:
        print("过切检查：无工序可处理")
        list_window.Close()
        return True

    # 启动加特林线程
    spammer = EnterSpammer(config['KEY_INTERVAL'])
    spammer.start()

    count_success = 0
    try:
        # 遍历工序执行过切检查
        for i, op in enumerate(operations):
            op_name = op.Name
            op_type = str(type(op))

            # 过滤钻孔类工序
            if any(x in op_type for x in ("Drill", "Hole", "Point")):
                continue

            msg = f"准备处理第 {i+1} 个工序: [{op_name}]"
            list_window.WriteLine(msg)  # 写入文件，保留
            print(msg)  # 控制台同步输出

            try:
                camSetup.GougeCheck([op], False)
                infoTool.DisplayCamObjectsDetails([op])

                success_msg = f"[OK] 成功: {op_name}"
                list_window.WriteLine(success_msg)
                print(success_msg)
                count_success += 1
            except Exception as e:
                err_msg = f"[FAIL] 失败 {op_name}: {str(e)}"
                list_window.WriteLine(err_msg)
                print(err_msg)

            time.sleep(config['BETWEEN_OPS'])

        # 保存过切检查后的部件（仅保留此文件）
        save_part(config['GAUGE_CHECK_SAVE_PATH'])

        # 写入完成信息
        finish_msg = f"过切检查完成。成功: {count_success} 个"
        list_window.WriteLine(finish_msg)
        print(finish_msg)

        return True
    finally:
        # 停止加特林线程
        spammer.stop()
        spammer.join()
        list_window.Close()  # 关闭文件写入

# ==================================================================================
# 5. NX工序参数导出器（完全移除ListingWindow）
# ==================================================================================
class NXOperationParamExporter:
    """NX工序参数导出器，用于批量扫描并汇总所有工序参数到单个JSON文件"""

    def __init__(self, session=None, work_part=None, config=None):
        self.theSession = session or NXOpen.Session.GetSession()
        self.theUFSession = NXOpen.UF.UFSession.GetUFSession()
        self.workPart = work_part or self.theSession.Parts.Work
        self.config = config or {}  # 移除lw相关初始化
        
        self.param_dictionary = self._get_param_dictionary()
        self.summary_data = {}
        self.batch_timestamp = time.strftime("%Y%m%d_%H%M%S", time.localtime())
        self.success_count = 0
        self.fail_count = 0

    def _get_param_dictionary(self):
        return {
            6:  "Display Tool Options",
            11: "Feed Engage",
            12: "Feed Retract",
            17: "Feed Cut",
            21: "Region Connection",
            24: "Corner Control Method",
            30: "Return Method",
            37: "Boundary Tolerances",
            42: "Min Clearance",
            45: "Start Method",
            49: "Gohome Method",
            55: "Motion Output Type",
            58: "NURBS Angular Tolerance",
            105: "Template Type",
            106: "Template Subtype",
            107: "Post Commands",
            116: "Suppressed",
            124: "Toolpath Time",
            125: "Toolpath Length",
            141: "Split Toolpath by Holder",
            142: "Toolpath Cutting Time",
            143: "Toolpath Cutting Length",
            148: "Postprocessor Cutting Time",
            153: "Template Class",
            154: "Template Subclass",
            221: "Clearance Plane Usage",
            1107: "Tool Adjust Reg Toggle",
            3010: "Hole Geometry",
            4005: "Spindle RPM",
            4013: "Spindle RPM Toggle",
            7210: "Last Tool Diameter",
            8212: "Tool Axis Tilt Data"
        }

    def init_environment(self):
        """初始化环境（控制台输出，替代窗口）"""
        print("\n正在启动工序参数导出程序...")

    def get_all_operations(self):
        """获取所有CAM工序（移除窗口输出，改为控制台）"""
        camSetup = self.workPart.CAMSetup
        if camSetup is None:
            print("\n【错误】当前部件中未检测到CAM加工环境！")
            return []

        opCollection = camSetup.CAMOperationCollection
        operations = [op for op in opCollection]
        if not operations:
            print("\n【错误】当前部件中未检测到任何工序！")
        else:
            print(f"\n检测到 {len(operations)} 个工序，开始批量扫描参数...")
        return operations

    def scan_operation_params(self, op):
        """扫描单个工序的参数"""
        obj_tag = op.Tag
        collected_params = []
        # 扫描参数范围：1-10000
        for index in range(1, 10000):
            val = None
            val_type = "Unknown"
            # Try Double
            if val is None:
                try:
                    val = self.theUFSession.Param.AskDoubleValue(obj_tag, index)
                    val_type = "Double"
                    if val == 0.0 and index not in self.param_dictionary:
                        val = None
                except:
                    pass
            # Try Int
            if val is None:
                try:
                    val = self.theUFSession.Param.AskIntValue(obj_tag, index)
                    val_type = "Int"
                    if val == 0 and index not in self.param_dictionary:
                        val = None
                except:
                    pass
            # Try String
            if val is None:
                try:
                    val = self.theUFSession.Param.AskStringValue(obj_tag, index)
                    val_type = "String"
                    if val == "":
                        val = None
                except:
                    pass
            # Try Tag
            if val is None:
                try:
                    val = self.theUFSession.Param.AskTagValue(obj_tag, index)
                    val_type = "Tag"
                    if val == NXOpen.Tag.Null:
                        val = None
                except:
                    pass

            if val is not None:
                display_name = self.param_dictionary.get(index, f"UNKNOWN_ID_{index}")
                collected_params.append({
                    "id": index,
                    "display_name": display_name,
                    "type": val_type,
                    "value": val
                })
        return collected_params

    def build_summary_data(self):
        """构建汇总数据"""
        nx_version = "NX"
        try:
            nx_version = self.theSession.EnvironmentInformation.Version
        except:
            pass

        self.summary_data = {
            "batch_meta": {
                "export_time": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
                "nx_version": nx_version,
                "part_name": self.workPart.Name if self.workPart else "未知部件",
                "total_operations": 0,
                "batch_timestamp": self.batch_timestamp,
                "success_operations": 0,
                "fail_operations": 0
            },
            "operations": []
        }

    def process_operations(self, operations):
        """处理所有工序的参数"""
        self.summary_data["batch_meta"]["total_operations"] = len(operations)
        for idx, op in enumerate(operations, 1):
            op_name = op.Name
            print(f"\n[{idx}/{len(operations)}] 正在处理工序：{op_name}")
            try:
                collected_params = self.scan_operation_params(op)
                self.summary_data["operations"].append({
                    "operation_name": op_name,
                    "operation_type": type(op).__name__,
                    "total_params": len(collected_params),
                    "parameters": collected_params,
                    "status": "success"
                })
                self.success_count += 1
                self.summary_data["batch_meta"]["success_operations"] = self.success_count
            except Exception as e:
                self.summary_data["operations"].append({
                    "operation_name": op_name,
                    "status": "failed",
                    "error_message": str(e)
                })
                self.fail_count += 1
                self.summary_data["batch_meta"]["fail_operations"] = self.fail_count
                print(f"    【处理失败】: {str(e)}")

    def save_summary_file(self):
        """保存JSON文件"""
        base_dir = self.config.get('JSON_EXPORT_BASE_DIR', os.getcwd())
        if not os.path.exists(base_dir):
            try:
                os.makedirs(base_dir)
            except OSError as e:
                print(f"\n【错误】无法创建目录 {base_dir}：{str(e)}")
                return False

        full_path = self.config.get('JSON_FULL_PATH', os.path.join(base_dir, "data.json"))
        try:
            with open(full_path, "w", encoding='utf-8') as f:
                json.dump(self.summary_data, f, indent=4, ensure_ascii=False)
            print(f"\n成功保存JSON文件: {full_path}")
            return True
        except Exception as e:
            print(f"\n【保存失败】: {str(e)}")
            return False

    def run(self):
        """执行参数导出流程"""
        self.init_environment()
        operations = self.get_all_operations()
        if not operations:
            return
        self.build_summary_data()
        self.process_operations(operations)
        self.save_summary_file()

# ==================================================================================
# 6. Excel 报告生成器 (数据处理逻辑)
# ==================================================================================
class ExcelReportGenerator:
    """处理文本和JSON数据，生成Excel报告"""
    
    @staticmethod
    def split_txt_by_generated_on(file_path):
        blocks = []
        current_block = []
        pattern = re.compile(r'\s*GENERATED\s+ON\s*', re.IGNORECASE)

        if not os.path.exists(file_path):
            print(f"ExcelGenerator: TXT文件不存在 {file_path}")
            return []

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line_stripped = line.rstrip('\n')
                    if pattern.search(line_stripped):
                        if current_block:
                            blocks.append(current_block)
                        current_block = [line_stripped]
                    else:
                        current_block.append(line_stripped)
                if current_block:
                    blocks.append(current_block)
            return blocks
        except Exception as e:
            print(f"ExcelGenerator: 读取文件错误 {e}")
            return []

    @staticmethod
    def extract_process_info(block):
        # 1. 提取工序名
        process_pattern = re.compile(r'(Operation Name)\s*[:：\s]\s*(.+)', re.IGNORECASE)
        # 2. 提取过切状态
        gouge_pattern = re.compile(r'(Gouge Check Status)\s*[:：\s]\s*([^；，。\n]+)', re.IGNORECASE)

        process_name = "未知工序"
        gouge_status = "None"

        for line in block:
            if process_name == "未知工序":
                process_match = process_pattern.search(line)
                if process_match:
                    process_name = process_match.group(2).strip()
            if gouge_status == "None":
                gouge_match = gouge_pattern.search(line)
                if gouge_match:
                    gouge_status = gouge_match.group(2).strip()
            if process_name != "未知工序" and gouge_status != "None":
                break
        return process_name, gouge_status

    @staticmethod
    def extract_toolpath_time_from_json(json_path):
        toolpath_times = []
        if not os.path.exists(json_path):
            print(f"ExcelGenerator: JSON文件不存在 {json_path}")
            return toolpath_times
        
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            operations = data.get('operations', [])
            for op in operations:
                toolpath_time = 0.0
                parameters = op.get('parameters', [])
                for param in parameters:
                    display_name = param.get('display_name', '').strip().lower()
                    if 'toolpath time' in display_name or '加工时间' in display_name:
                        try:
                            toolpath_time = float(param.get('value', 0.0))
                        except:
                            toolpath_time = 0.0
                        break
                toolpath_times.append(toolpath_time)
            print(f"从JSON中提取到 {len(toolpath_times)} 个Toolpath Time值")
            return toolpath_times
        except Exception as e:
            print(f"ExcelGenerator: JSON解析错误 {e}")
            return toolpath_times

    @staticmethod
    def write_to_excel(excel_path, process_names, gouge_statuses, toolpath_statuses, part_file_path):
        from openpyxl.utils import get_column_letter  # 局部导入避免影响顶部
        try:
            # 确保Excel目录存在（兜底处理）
            excel_dir = os.path.dirname(excel_path)
            if not os.path.exists(excel_dir):
                os.makedirs(excel_dir)

            if os.path.exists(excel_path):
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "CAM过切检查结果"
                ws.cell(row=1, column=1, value="文件路径")    # A1: 新增
                ws.cell(row=1, column=2, value="工序名")
                ws.cell(row=1, column=3, value="是否过切")
                ws.cell(row=1, column=4, value="是否生成刀路")
                print(f"已创建新Excel文件：{excel_path}")

            # 确保表头存在
            if ws.cell(row=1, column=1).value != "文件路径":
                ws.cell(row=1, column=1, value="文件路径")

            process_col = 2
            gouge_col = 3
            toolpath_col = 4
            start_row = 2

            # 写入数据
            max_len = max(len(process_names), len(gouge_statuses), len(toolpath_statuses))
            for idx in range(max_len):
                # 写入文件路径
                ws.cell(row=start_row + idx, column=1, value=part_file_path)
                
                if idx < len(process_names):
                    ws.cell(row=start_row + idx, column=process_col, value=process_names[idx])
                if idx < len(gouge_statuses):
                    ws.cell(row=start_row + idx, column=gouge_col, value=gouge_statuses[idx])
                if idx < len(toolpath_statuses):
                    ws.cell(row=start_row + idx, column=toolpath_col, value=toolpath_statuses[idx])
                else:
                    ws.cell(row=start_row + idx, column=toolpath_col, value="未知")

            # 自动调整列宽
            for column_cells in ws.columns:
                length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
                # 适当增加一点宽度作为缓冲，并设置上限防止过宽
                adjusted_width = min((length + 2) * 1.2, 50) 
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

            wb.save(excel_path)
            print(f"Excel报告生成成功！共写入 {max_len} 条数据。")
            print(f"路径: {excel_path}")
        except PermissionError:
            print("错误：Excel文件已被占用，请关闭后重试。")
        except Exception as e:
            print(f"Excel写入失败：{e}")
            traceback.print_exc()

    @classmethod
    def run_report_generation(cls, config):
        print("\n===== 开始生成Excel报告 =====")
        
        # 1. 处理TXT
        txt_blocks = cls.split_txt_by_generated_on(config['OUTPUT_PATH'])
        if not txt_blocks:
            print("提示: 未从TXT中提取到有效块，跳过")
            return

        process_info = [cls.extract_process_info(block) for block in txt_blocks]
        process_names = [info[0] for info in process_info]
        gouge_statuses = [info[1] for info in process_info]

        # 2. 处理JSON
        toolpath_times = cls.extract_toolpath_time_from_json(config['JSON_FULL_PATH'])

        # 3. 对齐数据 (以JSON为准)
        min_len = len(toolpath_times)
        if len(process_names) > min_len:
            print(f"ExcelGenerator: TXT数据({len(process_names)}) 多于 JSON数据({min_len})，正在截断...")
            process_names = process_names[:min_len]
            gouge_statuses = gouge_statuses[:min_len]
        elif len(toolpath_times) > len(process_names):
            print(f"ExcelGenerator: JSON数据({len(toolpath_times)}) 多于 TXT数据({len(process_names)})，部分行将为空")

        # 4. 生成刀路生成状态
        toolpath_statuses = ["是" if t > 0 else "否" for t in toolpath_times]

        # 5. 写入Excel
        cls.write_to_excel(config['EXCEL_REPORT_PATH'], process_names, gouge_statuses, toolpath_statuses, config['PART_PATH'])

# ==================================================================================
# 7. 主工作流（核心：仅接收prt路径和根文件夹路径，文件名称拼接PRT名称）
# ==================================================================================
def main(
    part_path: str,
    root_dir: str
):
    """整合主流程：刀轨 -> 过切 -> JSON -> Excel
    Args:
        part_path: PRT部件文件的完整路径
        root_dir: 根文件夹路径，脚本会在该路径下自动创建excel/txt/json/prt子文件夹
    """
    # 1. 固定配置（保留不变的参数，无需外部传入）
    fixed_config = {
        "AUTO_SAVE_TOOLPATH": False,
        "HEART_BEAT_EVERY": 10,
        "KEY_INTERVAL": 0.3,
        "BETWEEN_OPS": 0.5
    }

    # 2. 自动创建根目录（如果不存在）
    if not os.path.exists(root_dir):
        os.makedirs(root_dir)
        print(f"已创建根文件夹: {root_dir}")

    # 3. 自动创建子文件夹
    sub_dirs = ["excel", "txt", "json", "prt"]
    sub_dir_paths = {}
    for sub_dir in sub_dirs:
        sub_dir_path = os.path.join(root_dir, sub_dir)
        if not os.path.exists(sub_dir_path):
            os.makedirs(sub_dir_path)
            print(f"已创建子文件夹: {sub_dir_path}")
        sub_dir_paths[sub_dir] = sub_dir_path

    # 4. 提取PRT文件名（核心：只提取一次，全程复用）
    part_name = os.path.splitext(os.path.basename(part_path))[0]

    # 5. 自动生成各输出文件的完整路径（拼接PRT名称，保证一致性）
    dynamic_path_config = {
        "PART_PATH": part_path,
        # PRT子文件夹：PRT名 + _guoqie.prt
        "GAUGE_CHECK_SAVE_PATH": os.path.join(sub_dir_paths["prt"], f"{part_name}_guoqie.prt"),
        # TXT子文件夹：PRT名 + _Feature_Info.txt
        "OUTPUT_PATH": os.path.join(sub_dir_paths["txt"], f"{part_name}_Feature_Info.txt"),
        # JSON子文件夹：PRT名 + _data.json
        "JSON_EXPORT_BASE_DIR": sub_dir_paths["json"],
        "JSON_FULL_PATH": os.path.join(sub_dir_paths["json"], f"{part_name}_data.json"),
        # Excel子文件夹：PRT名 + _CAM过切检查.xlsx
        "EXCEL_REPORT_PATH": os.path.join(sub_dir_paths["excel"], f"{part_name}_CAM过切检查.xlsx")
    }

    # 6. 合并配置（动态路径 + 固定参数）
    config = {**dynamic_path_config, **fixed_config}

    # 7. 必传参数校验
    if not os.path.exists(part_path):
        raise ValueError(f"错误：PRT文件不存在 -> {part_path}")

    # 后续核心业务逻辑
    session = NXOpen.Session.GetSession()

    # 打开部件
    print(f"\n正在打开部件: {config['PART_PATH']}")
    try:
        base_part, load_status = session.Parts.OpenBaseDisplay(config['PART_PATH'])
    except Exception as e:
        print(f"打开部件失败: {e}")
        return False

    if not check_part_load_status(load_status):
        print("[FAIL] 打开部件失败")
        return False
    work_part = session.Parts.Work

    # 生成刀轨
    generator = ToolpathGeneratorMacro(session, work_part)
    generator.generate_all_toolpaths()
    generator.print_log("刀轨生成流程结束", "END")

    # 执行过切检查（生成TXT）
    print("\n===== 开始执行过切检查 =====")
    process_gauge_check(config)

    # 导出工序参数为JSON
    print("\n===== 开始导出工序参数为JSON =====")
    exporter = NXOperationParamExporter(session, work_part, config)
    exporter.run()

    # 生成Excel报告（已禁用：NX 内置 Python 没装 openpyxl）
    # ExcelReportGenerator.run_report_generation(config)

    print("\n===== 所有流程执行完成 =====")
    print(f"输出文件根目录: {root_dir}")
    return True

# ==================================================================================
# 主程序入口（测试用）
# ==================================================================================
if __name__ == "__main__":
    # 本地测试时手动传入两个参数：PRT路径 + 根文件夹路径
    try:
        part_path = r"c:\Users\Arno\Desktop\祝冕卜\过切检查\cpp_version\input\B1-01-M260040-P2.prt"
        root_dir = r"c:\Users\Arno\Desktop\祝冕卜\过切检查\python_test_output"

        main(part_path, root_dir)
    except Exception as e:
        print(f"\n[FAIL] 程序执行异常: {e}")
        traceback.print_exc()